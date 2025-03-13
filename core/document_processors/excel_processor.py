# core/document_processors/excel_processor.py
import os
from typing import Dict, Any, List
import logging
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

from .base_processor import BaseDocumentProcessor

logger = logging.getLogger(__name__)


class ExcelProcessor(BaseDocumentProcessor):
    """Excel文档处理器，用于处理Excel文件(xlsx, xls)"""

    def __init__(self, config=None):
        """
        初始化Excel处理器

        Args:
            config: 可选配置参数
        """
        super().__init__(config or {})
        self.sheet_separator = config.get("sheet_separator", "\n\n--- Sheet: {sheet_name} ---\n\n")
        self.max_rows = config.get("max_rows", None)  # 限制处理的最大行数
        self.max_cols = config.get("max_cols", None)  # 限制处理的最大列数
        self.extract_formulas = config.get("extract_formulas", True)  # 是否提取公式
        self.include_hidden_sheets = config.get("include_hidden_sheets", False)  # 是否包含隐藏工作表

        logger.info(f"初始化Excel处理器: max_rows={self.max_rows}, max_cols={self.max_cols}")

    def extract_text(self, file_path: str) -> str:
        """
        从Excel文件中提取文本内容

        Args:
            file_path: Excel文件路径

        Returns:
            提取的文本内容，包含各个工作表
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")

        try:
            # 读取Excel文件的所有工作表
            excel_file = pd.ExcelFile(file_path)
            all_sheets_text = []

            # 检查是否有隐藏的工作表
            hidden_sheets = set()
            if not self.include_hidden_sheets:
                try:
                    # 使用openpyxl检查隐藏状态
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    for sheet_name in wb.sheetnames:
                        if wb[sheet_name].sheet_state == 'hidden':
                            hidden_sheets.add(sheet_name)
                except Exception as e:
                    logger.warning(f"检查隐藏工作表时出错: {str(e)}")

            for sheet_name in excel_file.sheet_names:
                # 跳过隐藏工作表
                if sheet_name in hidden_sheets:
                    logger.debug(f"跳过隐藏工作表: {sheet_name}")
                    continue

                logger.debug(f"处理工作表: {sheet_name}")

                # 读取工作表数据
                df = pd.read_excel(
                    excel_file,
                    sheet_name=sheet_name,
                    nrows=self.max_rows,
                    usecols=range(self.max_cols) if self.max_cols else None
                )

                # 替换NaN为空字符串
                df = df.fillna("")

                # 将DataFrame转换为字符串表示
                sheet_text = df.to_string(index=False)

                # 添加工作表分隔符和内容
                all_sheets_text.append(
                    self.sheet_separator.format(sheet_name=sheet_name) + sheet_text
                )

            # 如果启用了公式提取，则添加公式信息
            if self.extract_formulas:
                formulas = self._extract_formulas(file_path)
                if formulas:
                    formula_text = "\n\n--- 电子表格公式 ---\n\n"
                    for f in formulas:
                        formula_text += f"Sheet '{f['sheet']}', Cell {f['cell']}: {f['formula']}\n"
                    all_sheets_text.append(formula_text)

            return "\n\n".join(all_sheets_text)

        except Exception as e:
            logger.error(f"提取Excel文本时出错: {str(e)}")
            raise

    def extract_metadata(self, file_path: str) -> Dict[str, Any]:
        """
        提取Excel文件的元数据

        Args:
            file_path: Excel文件路径

        Returns:
            包含文档元数据的字典
        """
        try:
            excel_file = pd.ExcelFile(file_path)

            # 提取工作簿元数据
            metadata = {}

            # 尝试使用openpyxl获取更多元数据
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)

                # 提取工作簿属性
                if wb.properties:
                    props = wb.properties
                    metadata.update({
                        "title": props.title,
                        "subject": props.subject,
                        "creator": props.creator,
                        "keywords": props.keywords,
                        "description": props.description,
                        "last_modified_by": props.lastModifiedBy,
                        "created": props.created.isoformat() if props.created else None,
                        "modified": props.modified.isoformat() if props.modified else None,
                        "category": props.category,
                        "content_status": props.contentStatus,
                    })
            except Exception as wb_err:
                logger.warning(f"提取工作簿属性时出错: {str(wb_err)}")

            # 提取各工作表的行列数
            sheets_info = []
            total_rows = 0

            for sheet_name in excel_file.sheet_names:
                # 计算每个工作表的行数和列数
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                rows, cols = df.shape
                total_rows += rows

                sheets_info.append({
                    "name": sheet_name,
                    "rows": rows,
                    "columns": cols,
                    "hidden": sheet_name in metadata.get("hidden_sheets", [])
                })

            # 构建元数据
            metadata.update({
                "file_name": os.path.basename(file_path),
                "file_extension": os.path.splitext(file_path)[1].lstrip(".").lower(),
                "file_size": os.path.getsize(file_path),
                "last_modified": datetime.fromtimestamp(
                    os.path.getmtime(file_path)
                ).isoformat(),
                "sheet_count": len(excel_file.sheet_names),
                "sheets": sheets_info,
                "total_rows": total_rows,
            })

            return {k: v for k, v in metadata.items() if v is not None}

        except Exception as e:
            logger.error(f"提取Excel元数据时出错: {str(e)}")
            raise

    @staticmethod
    def _extract_formulas(file_path: str) -> List[Dict[str, Any]]:
        """
        提取Excel文件中的公式

        Args:
            file_path: Excel文件路径

        Returns:
            包含公式信息的列表
        """
        formulas = []
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=False)

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                # 遍历所有单元格查找公式
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith('='):
                            formulas.append({
                                'sheet': sheet_name,
                                'cell': f"{get_column_letter(cell.column)}{cell.row}",
                                'formula': cell.value
                            })

            logger.debug(f"从Excel中提取了 {len(formulas)} 个公式")
            return formulas

        except Exception as e:
            logger.warning(f"提取Excel公式时出错: {str(e)}")
            return formulas

    @staticmethod
    def extract_pivot_tables(file_path: str) -> List[Dict[str, Any]]:
        """
        提取Excel文件中的数据透视表信息

        Args:
            file_path: Excel文件路径

        Returns:
            数据透视表信息列表
        """
        pivot_tables = []
        try:
            # 注意：openpyxl对数据透视表的支持有限
            # 这个方法可能需要在实际应用中使用其他库或方法来实现
            wb = openpyxl.load_workbook(file_path, read_only=True)

            # 检测每个工作表是否可能包含数据透视表
            # 这是一个简化的检测方法，不能准确识别所有数据透视表
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                # 查找可能的数据透视表标题
                for row in sheet.iter_rows(max_row=10):  # 只检查前10行
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and "汇总" in cell.value:
                            pivot_tables.append({
                                'sheet': sheet_name,
                                'possible_location': f"{get_column_letter(cell.column)}{cell.row}",
                                'title': cell.value
                            })
                            break

            return pivot_tables

        except Exception as e:
            logger.warning(f"提取数据透视表时出错: {str(e)}")
            return pivot_tables